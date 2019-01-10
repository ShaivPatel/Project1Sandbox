from openpyxl import load_workbook
from models.persons import Client, Employee
from models.emails import EmailRow, Email
import pandas
######################## Load Files ###############################
email_records_file_path =  "EmailRecordsPseudoData.xlsx"
sales_force_contacts_path =  "SalesForcePseudoData.csv"

salesforce_contacts_df = pandas.read_csv(sales_force_contacts_path)
email_records_ws = load_workbook(filename=email_records_file_path).worksheets[0]

###################################################################

######################## Assign Columns ###########################
date_col = "A"
id_col = "D"
email_address_col = "C"
type_col = "F"
subject_col = "E"
domain_col = "B"

company_name = "CompanyName"
company_domain = "Domain"
###################################################################

employees = {}
clients = {}





def parseSalesForce(df):
    clients = {}
    for row_index, row in df.iterrows():
        companyName = row[company_name]
        companyDomain = row[company_domain]
        clients[companyDomain] = Client(companyName, companyDomain)
    return clients


def parseEmailRecords(emails):
    row_cursor = 2
    email_buffer = []
    is_sender_employee = False
    while row_cursor<emails.max_row:
        e_row = EmailRow([cell.value for cell in list(emails[f"A{row_cursor}:F{row_cursor}"][0])])


        if e_row.type  == 'FROM':

            if len(email_buffer)>1:
                Email(email_buffer,employees,clients)
            email_buffer = []

            if e_row.domain != "macquarie.com" and e_row.domain not in clients:
                row_id = e_row.id
                while emails[f"{id_col}{row_cursor}"].value == row_id:
                    row_cursor += 1
            else:
                if e_row.domain == "macquarie.com":
                    is_sender_employee = True
                email_buffer.append(e_row)
                row_cursor += 1
        else:
            if is_sender_employee:
                if e_row.domain in clients:
                    email_buffer.append(e_row)
            else:
                if emails[f"{domain_col}{row_cursor}"].value == "macquarie.com":
                    email_buffer.append(e_row)
            row_cursor += 1




if __name__ == "__main__":

    clients = parseSalesForce(salesforce_contacts_df)

    parseEmailRecords(email_records_ws)


    print("===========================\nEmployee\n=============================")

    for employee in employees.values():
        print(f"===========================\n{employee}\n=============================")
        print("=============\nSent\n==============")
        for s in employee.sentEmails:
            print(s)
        print("=============\nReceived\n==============")
        for r in employee.receivedEmails:
            print(r)

    print("===========================\nClients\n=============================")

    for client in clients.values():
        print(f"===========================\n{client}\n=============================")
        print("=============\nSent\n==============")
        for s in client.sentEmails:
            print(s)
        print("=============\nReceived\n==============")
        for r in client.receivedEmails:
            print(r)